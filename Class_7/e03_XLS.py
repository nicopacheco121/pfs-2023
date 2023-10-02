"""

EXCEL


"""

### LECTURA ###

# Archivo CierreROFEX.xlsx

# Libreria que vamos a utilizar. Esto nos sirve para acceder a celdas, filas, columnas, etc
import openpyxl
import openpyxl.utils.cell

# abrimos el excel. El data_only=True sirve para no ver las formulas sino los valores
workbook = openpyxl.load_workbook('CierreROFEX.xlsx', data_only=True)
# Debugeamos y vemos el worksheets

print(workbook.sheetnames)
sheet = workbook.worksheets[0]  # puedo acceder a la hoja por nombre o por posicion

# Leo hoja
# Puedo acceder a los datos identificando las filas y columnas de diferente form: por letra y valor o por rango

title = [cell.value for cell in sheet[1]]  # con sheet[1] accedo a toda la fila 1 hasta la columna que tengo info
# el texto de cada celda esta en cell.value

print(title)

# Me quiero quedar con la columna de ajuste, pero no quiero dejar hardcodeado el numero de columna, sino que quiero buscarlo
column = sheet[openpyxl.utils.cell.get_column_letter(title.index("Ajuste") + 1)]
# title.index("Ajuste") me devuelve el numero de columna donde esta el texto "Ajuste"
# excel no tiene zero base index, por eso le sumo 1
# openpyxl.utils.cell.get_column_letter me devuelve la letra de la columna en base al numero de columna

# con sheet[letra] accedo a toda la columna

# Procesamos la informacion
# Leo las columna de posicion y ajuste y las guardo en un diccionario
dict = {"posicion": sheet[openpyxl.utils.cell.get_column_letter(title.index("Posicion") + 1)],
        "ajuste": sheet[openpyxl.utils.cell.get_column_letter(title.index("Ajuste") + 1)]}

# Imprimo la informacion, sin los titulos
for pos, ajuste in zip(dict["posicion"], dict["ajuste"]):
    if pos == "Posicion":
        continue
    print(pos.value, ajuste.value)

### ESCRITURA ###
newSheet = workbook.create_sheet("MiPrueba")  # Genero una nueva hoja
newSheet["A1"] = "matias"  # le asigno un valor a la celda A1

workbook.save('CierreROFEX.xlsx')  # le digo en que archivo quiero guardar los cambios

workbook.remove(newSheet)  # borro la hoja que cree

workbook.save('CierreROFEX.xlsx')  # guardo el archivo


# OTRO EJEMPLO QUE PUEDEN VER
workbook = openpyxl.load_workbook('dataset_reporte_covid_sitio_gobierno.xlsx', data_only=True)
sheet = workbook.worksheets[0]

title =[cell.value for cell in sheet[1]]
out = dict.fromkeys(["min_date","min","max_date","max"])
for row in sheet:
    if row[2].value == "ocupacion_de_camas_sistema_publico":
        if not(out["min"]) or out["min"]>=row[4].value:
            out["min_date"] = row[0].value
            out["min"] = row[4].value
        if not(out["max"]) or out["max"]<=row[4].value:
            out["max_date"] = row[0].value
            out["max"] = row[4].value

print(out)
